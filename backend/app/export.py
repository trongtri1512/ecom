"""Xuất danh sách quét ra Excel (xlsx) hoặc CSV."""
import csv
import io
from typing import List

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from .models import Scan

_HEADERS = ["Mã vận đơn", "ĐVVC", "NCC", "Ghi chú", "Thời gian quét", "Nguồn", "Số lần trùng"]


def to_import_xlsx(codes: List[str]) -> bytes:
    """Xuất file Excel ĐÚNG format import lên hệ thống imv.ops (1 cột 'Mã').

    Gồm 2 sheet như file mẫu: 'Dữ Liệu' (cột Mã) + 'Diễn Giải'.
    `codes`: danh sách mã vận đơn (thường là các đơn ĐÃ lấy hàng).
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Dữ Liệu"
    header_fill = PatternFill("solid", fgColor="808080")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    ws.column_dimensions["A"].width = 30
    ws["A1"].value = "Mã"
    ws["A1"].font = header_font
    ws["A1"].fill = header_fill
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    data_font = Font(name="Arial", size=14)
    for i, code in enumerate(codes, start=2):
        c = ws.cell(row=i, column=1, value=code)
        c.font = data_font
        c.alignment = Alignment(horizontal="left")

    ws2 = wb.create_sheet("Diễn Giải")
    hdrs = ["#", "Tên cột", "Diễn giải", "Giá trị", "Dữ liệu bắt buộc"]
    fill2 = PatternFill("solid", fgColor="7F7F7F")
    font2 = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    for col, h in enumerate(hdrs, 1):
        c = ws2.cell(row=1, column=col, value=h)
        c.font = font2
        c.fill = fill2
        c.alignment = Alignment(horizontal="center")
    note = ("Hỗ trợ Mã OR, mã OR đối tác, mã kiện hàng, mã vận đơn, mã vận đơn thu hồi\n\n"
            "Chú ý : nếu đơn hàng có nhiều kiện, bạn phải nhập vào mã kiện")
    for col, val in enumerate([1, "Mã", note, "Text", "YES"], 1):
        c = ws2.cell(row=2, column=col, value=val)
        c.font = Font(name="Calibri", size=11)
        c.alignment = Alignment(horizontal="center" if col in (1, 4, 5) else "left",
                                wrap_text=(col == 3))
    ws2.column_dimensions["B"].width = 15
    ws2.column_dimensions["C"].width = 60
    ws2.column_dimensions["D"].width = 12
    ws2.column_dimensions["E"].width = 20

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _row(s: Scan) -> list:
    return [
        s.code,
        s.carrier,
        s.supplier or "",
        s.note or "",
        s.scanned_at.strftime("%Y-%m-%d %H:%M:%S") if s.scanned_at else "",
        s.source_agent or "",
        s.dup_count or 0,
    ]


def to_xlsx(scans: List[Scan]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Scans"
    ws.append(_HEADERS)
    for s in scans:
        ws.append(_row(s))
    # Cột mã rộng hơn cho dễ đọc.
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["E"].width = 20
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def to_csv(scans: List[Scan]) -> bytes:
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(_HEADERS)
    for s in scans:
        writer.writerow(_row(s))
    # BOM để Excel mở tiếng Việt không lỗi font.
    return ("﻿" + buf.getvalue()).encode("utf-8")
